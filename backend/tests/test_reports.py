import csv
from collections.abc import AsyncIterator, Iterator
from datetime import date
from io import BytesIO, StringIO

import pytest
from httpx import ASGITransport, AsyncClient
from openpyxl import load_workbook
from sqlalchemy import create_engine, event
from sqlalchemy.orm import Session
from sqlalchemy.pool import StaticPool

from app.database.base import Base
from app.database.session import get_database_session
from app.main import app
from app.models import (
    Category,
    Expense,
    PaymentCycle,
    RecurringCostOpportunity,
    SpendingPriority,
    TransactionStatus,
)


@pytest.fixture
def session() -> Iterator[Session]:
    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )

    @event.listens_for(engine, "connect")
    def enable_foreign_keys(dbapi_connection, _connection_record) -> None:
        cursor = dbapi_connection.cursor()
        cursor.execute("PRAGMA foreign_keys=ON")
        cursor.close()

    Base.metadata.create_all(engine)
    with Session(engine, expire_on_commit=False) as database_session:
        yield database_session


def add_expense(
    session: Session,
    *,
    category: Category | None,
    transaction_date: date,
    amount: int,
    currency: str,
    status: TransactionStatus = TransactionStatus.CATEGORISED,
) -> None:
    session.add(
        Expense(
            transaction_date=transaction_date,
            description="Report transaction",
            normalised_description="REPORT TRANSACTION",
            amount=amount,
            currency=currency,
            category=category,
            status=status,
        )
    )


@pytest.mark.anyio
async def test_category_and_monthly_reports(session: Session) -> None:
    groceries = Category(name="Groceries")
    travel = Category(name="Travel")
    add_expense(
        session,
        category=groceries,
        transaction_date=date(2026, 1, 10),
        amount=-1000,
        currency="GBP",
    )
    add_expense(
        session,
        category=groceries,
        transaction_date=date(2026, 1, 20),
        amount=-2500,
        currency="GBP",
    )
    add_expense(
        session,
        category=travel,
        transaction_date=date(2026, 2, 1),
        amount=-5000,
        currency="GBP",
    )
    add_expense(
        session,
        category=travel,
        transaction_date=date(2026, 2, 2),
        amount=-2000,
        currency="EUR",
    )
    add_expense(
        session,
        category=None,
        transaction_date=date(2026, 2, 3),
        amount=9999,
        currency="GBP",
        status=TransactionStatus.NEEDS_REVIEW,
    )
    session.commit()

    async def override_database_session() -> AsyncIterator[Session]:
        yield session

    app.dependency_overrides[get_database_session] = override_database_session
    try:
        async with AsyncClient(
            transport=ASGITransport(app=app),
            base_url="http://testserver",
        ) as client:
            category_response = await client.get("/api/reports/category-totals?currency=GBP")
            monthly_response = await client.get(
                "/api/reports/monthly?date_from=2026-01-01&date_to=2026-02-28"
            )
    finally:
        app.dependency_overrides.clear()

    assert category_response.status_code == 200
    assert category_response.json()["items"] == [
        {
            "category_id": groceries.id,
            "category_code": None,
            "category_name": "Groceries",
            "currency": "GBP",
            "total_amount": 3500,
            "transaction_count": 2,
        },
        {
            "category_id": travel.id,
            "category_code": None,
            "category_name": "Travel",
            "currency": "GBP",
            "total_amount": 5000,
            "transaction_count": 1,
        },
    ]
    assert monthly_response.status_code == 200
    assert monthly_response.json()["items"] == [
        {
            "month": "2026-01",
            "currency": "GBP",
            "total_amount": 3500,
            "transaction_count": 2,
        },
        {
            "month": "2026-02",
            "currency": "EUR",
            "total_amount": 2000,
            "transaction_count": 1,
        },
        {
            "month": "2026-02",
            "currency": "GBP",
            "total_amount": 5000,
            "transaction_count": 1,
        },
    ]


@pytest.mark.anyio
async def test_reports_reject_reversed_date_range(session: Session) -> None:
    async def override_database_session() -> AsyncIterator[Session]:
        yield session

    app.dependency_overrides[get_database_session] = override_database_session
    try:
        async with AsyncClient(
            transport=ASGITransport(app=app),
            base_url="http://testserver",
        ) as client:
            response = await client.get(
                "/api/reports/monthly?date_from=2026-02-01&date_to=2026-01-01"
            )
    finally:
        app.dependency_overrides.clear()

    assert response.status_code == 422


@pytest.mark.anyio
async def test_exports_neutralise_spreadsheet_formulas(session: Session) -> None:
    category = Category(name="Shopping")
    descriptions = [
        '=HYPERLINK("https://example.invalid")',
        " +SUM(1,1)",
        "-2+3",
        "@SUM(1,1)",
        "Ordinary merchant",
    ]
    for description in descriptions:
        session.add(
            Expense(
                transaction_date=date(2026, 7, 24),
                description=description,
                normalised_description=description.upper(),
                amount=-100,
                currency="GBP",
                category=category,
                status=TransactionStatus.CATEGORISED,
            )
        )
    session.commit()

    async def override_database_session() -> AsyncIterator[Session]:
        yield session

    app.dependency_overrides[get_database_session] = override_database_session
    try:
        async with AsyncClient(
            transport=ASGITransport(app=app),
            base_url="http://testserver",
        ) as client:
            csv_response = await client.get("/api/reports/export?format=csv")
            xlsx_response = await client.get("/api/reports/export?format=xlsx")
    finally:
        app.dependency_overrides.clear()

    csv_rows = list(csv.DictReader(StringIO(csv_response.text)))
    assert [row["description"] for row in csv_rows] == [
        f"'{description}" if description != "Ordinary merchant" else description
        for description in descriptions
    ]
    assert csv_response.headers["cache-control"] == "no-store"
    assert csv_response.headers["x-content-type-options"] == "nosniff"

    workbook = load_workbook(BytesIO(xlsx_response.content), data_only=False)
    worksheet = workbook["Transactions"]
    exported_cells = [worksheet.cell(row=row, column=2) for row in range(2, 7)]
    assert [cell.value for cell in exported_cells] == [
        f"'{description}" if description != "Ordinary merchant" else description
        for description in descriptions
    ]
    assert all(cell.data_type == "s" for cell in exported_cells)
    assert xlsx_response.headers["cache-control"] == "no-store"
    assert xlsx_response.headers["x-content-type-options"] == "nosniff"


@pytest.mark.anyio
async def test_cash_flow_excludes_transfers_and_nets_refunds_and_income(
    session: Session,
) -> None:
    housing = Category(name="Housing")
    rent = Category(name="Rent", parent=housing)
    income = Category(name="Income")
    salary = Category(name="Salary", parent=income)
    transfers = Category(name="Transfers")
    own_account = Category(name="Own-account transfers", parent=transfers)
    savings = Category(
        name="Savings and investments",
        default_priority=SpendingPriority.TRANSFER,
    )
    for category, amount in (
        (rent, -10_000),
        (rent, 1_000),
        (salary, 20_000),
        (salary, -500),
        (own_account, -3_000),
        (savings, -2_000),
    ):
        add_expense(
            session,
            category=category,
            transaction_date=date(2026, 6, 15),
            amount=amount,
            currency="GBP",
        )
    session.commit()

    async def override_database_session() -> AsyncIterator[Session]:
        yield session

    app.dependency_overrides[get_database_session] = override_database_session
    try:
        async with AsyncClient(
            transport=ASGITransport(app=app),
            base_url="http://testserver",
        ) as client:
            categories = await client.get("/api/reports/category-totals?currency=GBP")
            monthly = await client.get("/api/reports/monthly?currency=GBP")
    finally:
        app.dependency_overrides.clear()

    assert categories.json()["items"] == [
        {
            "category_id": rent.id,
            "category_code": None,
            "category_name": "Rent",
            "currency": "GBP",
            "total_amount": 9_000,
            "transaction_count": 2,
        }
    ]
    assert monthly.json()["items"] == [
        {
            "month": "2026-06",
            "currency": "GBP",
            "total_amount": 9_000,
            "transaction_count": 2,
        }
    ]


@pytest.mark.anyio
async def test_payment_period_report_uses_cycle_and_priority_semantics(
    session: Session,
) -> None:
    cycle = PaymentCycle(
        name="Universal Credit",
        start_date=date(2026, 6, 25),
        end_date=date(2026, 8, 1),
        next_payment_date=date(2026, 7, 25),
        expected_income_amount=80_000,
        currency="GBP",
        opening_balance=20_000,
    )
    housing = Category(name="Housing", default_priority=SpendingPriority.PROTECTED)
    income = Category(name="Income")
    benefits = Category(name="Benefits", parent=income)
    optional = Category(name="Eating out", default_priority=SpendingPriority.OPTIONAL)
    session.add_all(
        [
            Expense(
                transaction_date=date(2026, 6, 26),
                description="Benefit payment",
                normalised_description="BENEFIT PAYMENT",
                amount=80_000,
                currency="GBP",
                category=benefits,
                payment_cycle=cycle,
                status=TransactionStatus.CATEGORISED,
            ),
            Expense(
                transaction_date=date(2026, 7, 1),
                description="Rent",
                normalised_description="RENT",
                amount=-50_000,
                currency="GBP",
                category=housing,
                payment_cycle=cycle,
                status=TransactionStatus.CATEGORISED,
            ),
            Expense(
                transaction_date=date(2026, 7, 2),
                description="Cafe",
                normalised_description="CAFE",
                amount=-1_000,
                currency="GBP",
                category=optional,
                payment_cycle=cycle,
                status=TransactionStatus.CATEGORISED,
            ),
        ]
    )
    session.commit()

    async def override_database_session() -> AsyncIterator[Session]:
        yield session

    app.dependency_overrides[get_database_session] = override_database_session
    try:
        async with AsyncClient(
            transport=ASGITransport(app=app),
            base_url="http://testserver",
        ) as client:
            response = await client.get("/api/reports/payment-periods?currency=GBP")
    finally:
        app.dependency_overrides.clear()

    assert response.status_code == 200
    assert response.json()["items"] == [
        {
            "payment_cycle_id": cycle.id,
            "name": "Universal Credit",
            "start_date": "2026-06-25",
            "end_date": "2026-08-01",
            "next_payment_date": "2026-07-25",
            "currency": "GBP",
            "status": "planned",
            "income": 80_000,
            "spending": 51_000,
            "net": 29_000,
            "transaction_count": 3,
            "protected_spending": 50_000,
            "essential_spending": 0,
            "adjustable_spending": 0,
            "optional_spending": 1_000,
            "irregular_essential_spending": 0,
        }
    ]


@pytest.mark.anyio
async def test_recurring_opportunity_stores_user_backed_saving(
    session: Session,
) -> None:
    subscriptions = Category(
        name="Subscriptions",
        default_priority=SpendingPriority.OPTIONAL,
    )
    for transaction_date in (
        date(2026, 4, 1),
        date(2026, 5, 1),
        date(2026, 6, 1),
    ):
        session.add(
            Expense(
                transaction_date=transaction_date,
                description="Video streaming",
                normalised_description="VIDEO STREAMING",
                amount=-1_800,
                currency="GBP",
                category=subscriptions,
                status=TransactionStatus.CATEGORISED,
            )
        )
    session.commit()

    async def override_database_session() -> AsyncIterator[Session]:
        yield session

    app.dependency_overrides[get_database_session] = override_database_session
    try:
        async with AsyncClient(
            transport=ASGITransport(app=app),
            base_url="http://testserver",
        ) as client:
            initial = await client.get("/api/reports/recurring-opportunities")
            saved = await client.put(
                "/api/reports/recurring-opportunities",
                json={
                    "description": "VIDEO STREAMING",
                    "currency": "GBP",
                    "current_monthly_cost": 1800,
                    "replacement_monthly_cost": 700,
                    "one_off_switching_cost": 0,
                    "difficulty": "easy",
                    "decision": "planned",
                },
            )
            same_identity = await client.put(
                "/api/reports/recurring-opportunities",
                json={
                    "description": "  video streaming ",
                    "currency": "gbp",
                    "current_monthly_cost": 1800,
                    "replacement_monthly_cost": 600,
                    "one_off_switching_cost": 0,
                    "difficulty": "easy",
                    "decision": "planned",
                },
            )
            ranked = await client.get("/api/reports/recurring-opportunities")
    finally:
        app.dependency_overrides.clear()

    assert initial.status_code == 200
    assert initial.json()["items"][0]["monthly_saving"] is None
    assert saved.status_code == 200
    assert saved.json()["monthly_saving"] == 1100
    assert saved.json()["identity_key"] == "description:video streaming"
    assert same_identity.status_code == 200
    assert same_identity.json()["opportunity_id"] == saved.json()["opportunity_id"]
    assert len(session.query(RecurringCostOpportunity).all()) == 1
    opportunity = ranked.json()["items"][0]
    assert opportunity["identity_key"] == "description:video streaming"
    assert opportunity["monthly_saving"] == 1200
    assert opportunity["first_year_saving"] == 14_400
    assert opportunity["difficulty"] == "easy"
    assert opportunity["decision"] == "planned"
