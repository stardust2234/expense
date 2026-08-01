import csv
from datetime import date, datetime
from hashlib import sha256
from io import BytesIO, StringIO
from pathlib import Path
from typing import Annotated
from zipfile import BadZipFile

from fastapi import (
    APIRouter,
    Depends,
    File,
    Form,
    HTTPException,
    Query,
    Response,
    UploadFile,
    status,
)
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from pypdf import PdfReader
from pypdf.errors import PdfReadError
from sqlalchemy.orm import Session, sessionmaker

from app.database.session import get_database_session
from app.schemas.imports import ImportBatchItem, ImportBatchListResponse
from app.services.csv_importer import CsvImportError, import_csv
from app.services.import_batch_service import (
    ACTIVE_IMPORT_STATUSES,
    ImportBatchConflictError,
    ImportBatchNotFoundError,
    batch_counts,
    delete_import_batch,
    find_duplicate_batch,
    get_import_batch,
    list_import_batches,
)
from app.services.import_job_service import enqueue_import_job, queue_import_batch

MAX_STATEMENT_BYTES = 10 * 1024 * 1024
router = APIRouter(prefix="/imports", tags=["imports"])
DatabaseSession = Annotated[Session, Depends(get_database_session)]


def _batch_item(batch) -> ImportBatchItem:
    normalised, failed, duplicates, categorised, needs_review = batch_counts(batch)
    return ImportBatchItem(
        id=batch.id,
        source_filename=batch.source_filename,
        source_type=batch.source_type,
        content_sha256=batch.content_sha256,
        default_currency=batch.default_currency,
        total_rows=batch.total_rows,
        normalised_rows=normalised,
        failed_rows=failed,
        duplicate_rows=duplicates,
        categorised_rows=categorised,
        needs_review_rows=needs_review,
        status=batch.processing_status,
        processing_error=batch.processing_error,
        imported_at=batch.imported_at,
        processing_started_at=batch.processing_started_at,
        processing_completed_at=batch.processing_completed_at,
    )


@router.get("", response_model=ImportBatchListResponse)
async def import_history(
    session: DatabaseSession,
    limit: Annotated[int, Query(ge=1, le=100)] = 25,
    offset: Annotated[int, Query(ge=0)] = 0,
) -> ImportBatchListResponse:
    page = list_import_batches(session, limit=limit, offset=offset)
    return ImportBatchListResponse(
        items=[_batch_item(batch) for batch in page.items],
        total=page.total,
        limit=limit,
        offset=offset,
    )


@router.get("/{batch_id}", response_model=ImportBatchItem)
async def import_detail(batch_id: int, session: DatabaseSession) -> ImportBatchItem:
    try:
        return _batch_item(get_import_batch(session, batch_id=batch_id))
    except ImportBatchNotFoundError as error:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(error)) from error


@router.delete("/{batch_id}", status_code=status.HTTP_204_NO_CONTENT)
async def remove_import(batch_id: int, session: DatabaseSession) -> Response:
    try:
        delete_import_batch(session, batch_id=batch_id)
    except ImportBatchNotFoundError as error:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(error)) from error
    except ImportBatchConflictError as error:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail=str(error)) from error
    return Response(status_code=status.HTTP_204_NO_CONTENT)


@router.post(
    "/{batch_id}/retry",
    response_model=ImportBatchItem,
    status_code=status.HTTP_202_ACCEPTED,
)
async def retry_import(
    batch_id: int,
    session: DatabaseSession,
) -> ImportBatchItem:
    try:
        batch = get_import_batch(session, batch_id=batch_id)
        _, failed, _, _, _ = batch_counts(batch)
        if batch.processing_status in ACTIVE_IMPORT_STATUSES:
            raise ImportBatchConflictError(
                f"Import batch {batch_id} is already {batch.processing_status}"
            )
        if not failed and batch.processing_status != "failed":
            raise ImportBatchConflictError(f"Import batch {batch_id} has no failed rows to retry")
    except ImportBatchNotFoundError as error:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(error)) from error
    except ImportBatchConflictError as error:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail=str(error)) from error

    queue_import_batch(session, batch)
    worker_sessions = sessionmaker(
        bind=session.get_bind(),
        autoflush=False,
        expire_on_commit=False,
    )
    queued_batch_id = batch.id
    item = _batch_item(batch)
    session.rollback()
    enqueue_import_job(
        queued_batch_id,
        retry_failed=True,
        session_factory=worker_sessions,
    )
    return item


@router.post("/file", response_model=ImportBatchItem, status_code=status.HTTP_202_ACCEPTED)
async def upload_statement(
    session: DatabaseSession,
    file: Annotated[UploadFile, File()],
    default_currency: Annotated[str | None, Form(min_length=3, max_length=3)] = None,
) -> ImportBatchItem:
    return await _process_upload(
        session,
        file,
        default_currency,
    )


async def _process_upload(
    session: Session,
    file: UploadFile,
    default_currency: str | None,
) -> ImportBatchItem:
    filename = file.filename or "transactions.csv"
    suffix = Path(filename).suffix.lower()
    allowed_suffixes = {".csv", ".xlsx", ".pdf"}
    if suffix not in allowed_suffixes:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_CONTENT,
            detail=f"Supported file types: {', '.join(sorted(allowed_suffixes))}",
        )
    content = await file.read(MAX_STATEMENT_BYTES + 1)
    if len(content) > MAX_STATEMENT_BYTES:
        raise HTTPException(
            status_code=status.HTTP_413_CONTENT_TOO_LARGE,
            detail="Statement file must not exceed 10 MiB",
        )
    content_digest = sha256(content).hexdigest()
    duplicate = find_duplicate_batch(session, content_sha256=content_digest)
    if duplicate is not None:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=(
                f"This file was already imported as batch {duplicate.id} "
                f"({duplicate.source_filename})"
            ),
        )

    try:
        text = _statement_to_csv(content, suffix)
    except (UnicodeDecodeError, ValueError) as error:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_CONTENT,
            detail=str(error),
        ) from error

    try:
        batch = import_csv(
            session,
            source_filename=filename,
            stream=StringIO(text, newline=""),
            source_type=suffix.removeprefix("."),
            content_sha256=content_digest,
            default_currency=default_currency.upper() if default_currency else None,
        )
    except CsvImportError as error:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_CONTENT,
            detail=str(error),
        ) from error

    queue_import_batch(session, batch)
    worker_sessions = sessionmaker(
        bind=session.get_bind(),
        autoflush=False,
        expire_on_commit=False,
    )
    queued_batch_id = batch.id
    item = _batch_item(batch)
    session.rollback()
    enqueue_import_job(
        queued_batch_id,
        session_factory=worker_sessions,
    )
    return item


def _statement_to_csv(content: bytes, suffix: str) -> str:
    if suffix == ".csv":
        try:
            return content.decode("utf-8-sig")
        except UnicodeDecodeError as error:
            raise ValueError("CSV file must use UTF-8 encoding") from error

    if suffix == ".xlsx":
        try:
            workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        except (BadZipFile, InvalidFileException) as error:
            raise ValueError("Excel file is invalid or unreadable") from error
        worksheet = workbook.active
        output = StringIO()
        writer = csv.writer(output)
        for row in worksheet.iter_rows(values_only=True):
            writer.writerow([_excel_value(value) for value in row])
        return output.getvalue()

    try:
        reader = PdfReader(BytesIO(content))
    except PdfReadError as error:
        raise ValueError("PDF file is invalid or unreadable") from error
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    comma_lines = [line for line in lines if "," in line]
    tab_lines = [line for line in lines if "\t" in line]
    table_lines = comma_lines if len(comma_lines) >= 2 else tab_lines
    if len(table_lines) < 2:
        raise ValueError(
            "PDF does not contain a detectable CSV or tab-delimited transaction table; "
            "convert scanned statements to CSV or Excel first"
        )
    if table_lines is tab_lines:
        output = StringIO()
        writer = csv.writer(output)
        for line in table_lines:
            writer.writerow(next(csv.reader([line], delimiter="\t")))
        return output.getvalue()
    return "\n".join(table_lines)


def _excel_value(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value
